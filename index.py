import sys
import os
import time

import pandas as pd
import numpy as np
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.lib.units import mm

# =========================
# CONFIG
# =========================

COLUNAS = [
    "Nota Fiscal", "Modelo", "Data", "CNPJ/CPF",
    "CCE", "UF", "CFOP", "CST / Mercadoria",
    "Categoria", "Ct Sefaz", "Ct Contrib",
    "Dif CT", "Valor Produto", "Dif. Icms",
]

GRUPOS = [
    ("Documento", 3),
    ("Destino", 4),
    ("Produto", 2),
    ("Tributação", 5),
]

COR_AZUL = colors.HexColor("#4472C4")
COR_AZUL_CLARO = colors.HexColor("#D9E2F3")

STYLE_TITULO = ParagraphStyle(
    "titulo", fontSize=11, leading=13, alignment=TA_CENTER,
    fontName="Helvetica-Bold",
)
STYLE_INFO = ParagraphStyle(
    "info", fontSize=8, leading=10, alignment=TA_LEFT,
)
STYLE_HEADER = ParagraphStyle(
    "header", fontSize=7, leading=9, alignment=TA_CENTER,
    textColor=colors.whitesmoke, fontName="Helvetica-Bold",
)
STYLE_CELL = ParagraphStyle(
    "cell", fontSize=6.5, leading=8, alignment=TA_LEFT,
)


def _safe(val) -> str:
    s = str(val)
    if s in ("nan", "NaT", "None"):
        return ""
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def extrair_header(caminho: str) -> dict:
    """Extrai informações do cabeçalho institucional da primeira aba."""
    df = pd.read_excel(caminho, sheet_name=0, header=None, nrows=6)
    info = {}

    row0 = df.iloc[0].fillna("").astype(str)
    info["orgao"] = row0.iloc[0] if row0.iloc[0] != "" else "ESTADO DE GOIÁS SECRETARIA DA FAZENDA"
    titulo_parts = [v for v in row0 if "Diverg" in v]
    info["titulo"] = titulo_parts[0] if titulo_parts else "Divergências de Carga Tributária Informada e Calculada - Nota Fiscal"

    row1 = df.iloc[1].fillna("").astype(str)
    parts = [v.strip() for v in row1 if v.strip() and v.strip() != "nan"]
    info["empresa_parts"] = parts

    row2 = df.iloc[2].fillna("").astype(str)
    datas = []
    for v in row2:
        v = v.strip()
        if v and v not in ("Período :", "a", "nan", "NaT", ""):
            try:
                dt = pd.to_datetime(v)
                datas.append(dt.strftime("%d/%m/%Y"))
            except Exception:
                pass
    if len(datas) >= 2:
        info["periodo"] = f"Período :   {datas[0]}    a    {datas[1]}"
    else:
        info["periodo"] = "Período : -"

    # Extrair logo da primeira aba
    from openpyxl import load_workbook
    wb = load_workbook(caminho)
    ws = wb[wb.sheetnames[0]]
    logo_path = os.path.join(os.path.dirname(caminho) or ".", "logo_sefaz.png")
    info["logo"] = None
    for img in ws._images:
        with open(logo_path, "wb") as f:
            f.write(img._data())
        info["logo"] = logo_path
        break
    wb.close()

    # Extrair assinatura da última aba (Auditores Fiscais)
    sheets = pd.read_excel(caminho, sheet_name=None, header=None, engine="openpyxl")
    last_df = list(sheets.values())[-1]
    info["assinatura"] = []
    for _, row in last_df.iterrows():
        vals = row.fillna("").astype(str)
        first = vals.iloc[0].strip()
        if first.startswith("Auditores"):
            labels = [v.strip() for v in vals if v.strip() and v.strip() != "nan"]
            info["assinatura"].append(labels)
        elif info["assinatura"]:
            vals_clean = [v.strip() for v in vals if v.strip() and v.strip() != "nan"]
            if vals_clean:
                info["assinatura"].append(vals_clean)

    return info


import re

MARCA_MES = "__MES__"
MARCA_SUBTOTAL = "__SUB__"
_RE_MES = re.compile(r"^\d{1,2}/\d{4}$")

HEADER_KEYS = {
    "Nota Fiscal": "Nota Fiscal",
    "Modelo": "Modelo",
    "Data": "Data",
    "CNPJ/CPF": "CNPJ/CPF",
    "CCE": "CCE",
    "UF": "UF",
    "CFOP": "CFOP",
    "CST": "CST / Mercadoria",
    "Mercadoria": "Mercadoria",
    "Categoria": "Categoria",
    "Ct Sefaz": "Ct Sefaz",
    "Ct Contrib": "Ct Contrib",
    "Dif CT": "Dif CT",
    "Valor Produto": "Valor Produto",
    "Dif. Icms": "Dif. Icms",
}


def _find_mes(row_values) -> str | None:
    for v in row_values:
        s = str(v).strip()
        if _RE_MES.match(s):
            return s
    return None


def _find_subtotal(row_values) -> str | None:
    strs = [str(v).strip() for v in row_values]
    non_empty = [s for s in strs if s not in ("", "nan", "NaT", "None")]
    if len(non_empty) == 1:
        try:
            float(non_empty[0])
            return non_empty[0]
        except ValueError:
            pass
    if len(non_empty) == 2:
        mes_val, num_val = non_empty
        if _RE_MES.match(mes_val):
            try:
                float(num_val)
                return None
            except ValueError:
                pass
    return None


def _find_mes_with_total(row_values) -> tuple[str, str] | None:
    """Detecta linha de resumo: mês + valor (ex: '06/2016' + '1235.51')."""
    strs = [str(v).strip() for v in row_values]
    non_empty = [s for s in strs if s not in ("", "nan", "NaT", "None")]
    if len(non_empty) == 2:
        mes_val, num_val = non_empty
        if _RE_MES.match(mes_val):
            try:
                float(num_val)
                return (mes_val, num_val)
            except ValueError:
                pass
    return None


def _find_col_map(row_values) -> dict[str, int] | None:
    """Identifica posições das colunas a partir da linha de cabeçalho."""
    strs = {str(v).strip(): i for i, v in enumerate(row_values)
            if str(v).strip() not in ("", "nan", "NaT", "None")}
    if "Nota Fiscal" not in strs:
        return None
    col_map = {}
    for key, target in HEADER_KEYS.items():
        if key in strs:
            col_map[target] = strs[key]
    return col_map if len(col_map) >= 8 else None


def _val(row_values, idx: int):
    if idx is None or idx >= len(row_values):
        return ""
    v = row_values[idx]
    s = str(v).strip()
    if s in ("nan", "NaT", "None", ""):
        return ""
    return v


def _calibrate_col_map(col_map: dict[str, int], header_row, data_row) -> dict[str, int]:
    """Ajusta o mapa combinando header e primeira linha de dados.
    Só aplica offset +1 se: header[idx+1] é vazio E dados[idx] é vazio E dados[idx+1] não é vazio."""
    adjusted = {}
    n = len(header_row)
    nd = len(data_row)
    mapped_positions = set(col_map.values())

    for col_name, idx in col_map.items():
        if idx + 1 < n and idx + 1 < nd:
            hdr_next = str(header_row[idx + 1]).strip()
            hdr_next_empty = hdr_next in ("nan", "NaT", "None", "")
            next_mapped = (idx + 1) in mapped_positions

            dat_here = str(data_row[idx]).strip() if idx < nd else ""
            dat_next = str(data_row[idx + 1]).strip() if idx + 1 < nd else ""
            here_empty = dat_here in ("nan", "NaT", "None", "")
            next_has_val = dat_next not in ("nan", "NaT", "None", "")

            if hdr_next_empty and not next_mapped and here_empty and next_has_val:
                adjusted[col_name] = idx + 1
            else:
                adjusted[col_name] = idx
        else:
            adjusted[col_name] = idx
    return adjusted


def _extract_row(row_values, col_map: dict[str, int]) -> list:
    result = []
    for col_name in COLUNAS:
        if col_name == "CST / Mercadoria":
            idx = col_map.get("CST / Mercadoria", col_map.get("Mercadoria"))
        else:
            idx = col_map.get(col_name)
        result.append(_val(row_values, idx))
    return result


def ler_dados(caminho: str) -> pd.DataFrame:
    """Lê todas as abas usando mapa de colunas por cabeçalho."""
    if not os.path.isfile(caminho):
        print(f"ERRO: arquivo '{caminho}' nao encontrado.")
        sys.exit(1)

    print("  Carregando abas...")
    t0 = time.time()
    sheets: dict[str, pd.DataFrame] = pd.read_excel(
        caminho, sheet_name=None, header=None, engine="openpyxl",
    )
    print(f"  {len(sheets)} abas lidas em {time.time() - t0:.1f}s")

    n_target = len(COLUNAS)
    all_rows: list[list] = []

    for sheet_name, df in sheets.items():
        if df is None or df.empty or df.shape[1] == 0:
            continue

        raw_col_map = None
        header_vals = None
        col_map = None
        has_data = False
        sheet_rows: list[list] = []

        for _, row in df.iterrows():
            vals = row.values
            first = str(vals[0]).strip()

            if first in ("Documento", "Período :"):
                continue
            if first.startswith("ESTADO") or first.startswith("Raz"):
                continue
            if first.startswith("Auditores"):
                continue

            # "Resumo:" com total geral
            if first.startswith("Resumo"):
                strs = [str(v).strip() for v in vals]
                nums = []
                for s in strs:
                    if s not in ("", "nan", "NaT", "None", "Resumo:"):
                        try:
                            nums.append(float(s))
                        except ValueError:
                            pass
                if nums:
                    row_data = [MARCA_MES + "RESUMO"] + [""] * (n_target - 2) + [str(nums[0])]
                    sheet_rows.append(row_data)
                continue

            # "Referência\n03/2016" → extrair mês
            if first.startswith("Refer"):
                m_ref = re.search(r"\d{1,2}/\d{4}", first)
                if m_ref:
                    mes_ref = m_ref.group()
                    strs = [str(v).strip() for v in vals]
                    nums = [s for s in strs if s not in ("", "nan", "NaT", "None")
                            and not s.startswith("Refer") and s != mes_ref]
                    total_ref = ""
                    for s in nums:
                        try:
                            total_ref = str(float(s))
                            break
                        except ValueError:
                            pass
                    row_data = [MARCA_MES + mes_ref] + [""] * (n_target - 2) + [total_ref]
                    sheet_rows.append(row_data)
                continue

            if first == "Nota Fiscal":
                new_map = _find_col_map(vals)
                if new_map:
                    raw_col_map = new_map
                    header_vals = vals
                    col_map = None
                continue

            mes_total = _find_mes_with_total(vals)
            if mes_total:
                m, t = mes_total
                row_data = [MARCA_MES + m] + [""] * (n_target - 2) + [t]
                sheet_rows.append(row_data)
                continue

            mes = _find_mes(vals)
            if mes:
                sheet_rows.append([MARCA_MES + mes] + [""] * (n_target - 1))
                continue

            sub = _find_subtotal(vals)
            if sub:
                sheet_rows.append([MARCA_SUBTOTAL] + [""] * (n_target - 2) + [sub])
                continue

            if first in ("", "nan", "NaT", "None"):
                continue

            try:
                int(float(first))
            except (ValueError, OverflowError):
                continue

            if raw_col_map is None:
                continue

            if col_map is None:
                col_map = _calibrate_col_map(raw_col_map, header_vals, vals)

            extracted = _extract_row(vals, col_map)
            sheet_rows.append(extracted)
            has_data = True

        if has_data:
            all_rows.extend(sheet_rows)
        elif sheet_rows:
            is_summary = any(MARCA_SUBTOTAL in str(r[0]) for r in sheet_rows)
            if not is_summary:
                all_rows.extend(sheet_rows)

    if not all_rows:
        print("ERRO: nenhum dado encontrado.")
        sys.exit(1)

    # Garantir que meses inline sem total herdem o total do resumo
    month_totals: dict[str, str] = {}
    for row in all_rows:
        first = str(row[0])
        if first.startswith(MARCA_MES):
            mes = first.replace(MARCA_MES, "")
            last_val = str(row[-1]).strip()
            if last_val not in ("", "nan", "NaT", "None"):
                month_totals[mes] = last_val

    patched = 0
    for row in all_rows:
        first = str(row[0])
        if first.startswith(MARCA_MES):
            mes = first.replace(MARCA_MES, "")
            last_val = str(row[-1]).strip()
            if last_val in ("", "nan", "NaT", "None") and mes in month_totals:
                row[-1] = month_totals[mes]
                patched += 1

    if patched:
        print(f"  {patched} separadores de mes receberam total")

    df_final = pd.DataFrame(all_rows, columns=COLUNAS)
    return df_final


COR_MES = colors.HexColor("#E2EFDA")
STYLE_MES = ParagraphStyle(
    "mes", fontSize=8, leading=10, alignment=TA_LEFT,
    fontName="Helvetica-Bold",
)


def gerar_pdf(df: pd.DataFrame, saida: str, header_info: dict) -> None:
    df_str = df.fillna("").astype(str).replace("nan", "")
    n_rows, n_cols = df_str.shape
    print(f"  Montando PDF ({n_rows} linhas x {n_cols} colunas)...")
    t0 = time.time()

    from reportlab.platypus import Image as RLImage

    page_width, _ = landscape(A3)
    MARGIN = 15
    usable_w = page_width - (MARGIN * 2)

    elements = []

    # --- Linha 1: Logo + Órgão (esquerda) | Título (direita) ---
    STYLE_ORGAO = ParagraphStyle(
        "orgao", fontSize=8, leading=10, alignment=TA_LEFT,
        fontName="Helvetica-Bold",
    )
    STYLE_TITULO_R = ParagraphStyle(
        "titulo_r", fontSize=11, leading=14, alignment=TA_LEFT,
        fontName="Helvetica-Bold", textColor=colors.HexColor("#1F4E79"),
    )

    logo_path = header_info.get("logo")
    if logo_path and os.path.isfile(logo_path):
        logo_img = RLImage(logo_path, width=35, height=45)
    else:
        logo_img = ""

    orgao_text = Paragraph(_safe(header_info.get("orgao", "")), STYLE_ORGAO)
    titulo_text = Paragraph(
        f"<b><i>{_safe(header_info.get('titulo', ''))}</i></b>",
        STYLE_TITULO_R,
    )

    header_table = Table(
        [[logo_img, orgao_text, titulo_text]],
        colWidths=[40, usable_w * 0.25, usable_w * 0.70],
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 2 * mm))

    # --- Linha 2: Dados da empresa ---
    STYLE_EMP = ParagraphStyle("emp", fontSize=7.5, leading=9, alignment=TA_LEFT)
    empresa_parts = header_info.get("empresa_parts", [])
    empresa_text = "    ".join(_safe(p) for p in empresa_parts)
    elements.append(Paragraph(empresa_text, STYLE_EMP))
    elements.append(Spacer(1, 1.5 * mm))

    # --- Linha 3: Período ---
    STYLE_PER = ParagraphStyle(
        "per", fontSize=8, leading=10, alignment=TA_LEFT,
        fontName="Helvetica-Bold",
    )
    elements.append(Paragraph(_safe(header_info.get("periodo", "")), STYLE_PER))
    elements.append(Spacer(1, 3 * mm))

    # --- Linha de grupo ---
    grupo_row = []
    for nome, span in GRUPOS:
        grupo_row.append(Paragraph(f"<b>{_safe(nome)}</b>", STYLE_HEADER))
        for _ in range(span - 1):
            grupo_row.append("")
    while len(grupo_row) < n_cols:
        grupo_row.append("")
    grupo_row = grupo_row[:n_cols]

    # --- Linha de colunas ---
    col_row = [
        Paragraph(f"<b>{_safe(c)}</b>", STYLE_HEADER)
        for c in df_str.columns
    ]

    # --- Dados (preservando meses e subtotais) ---
    values = df_str.values.tolist()
    data = [grupo_row, col_row]
    mes_row_indices = []
    sub_row_indices = []

    for row in values:
        first = str(row[0])
        if first.startswith(MARCA_MES):
            mes_label = first.replace(MARCA_MES, "")
            mes_cells = [Paragraph(f"<b>{_safe(mes_label)}</b>", STYLE_MES)]
            mes_cells += [""] * (n_cols - 2)
            total_val = _safe(str(row[-1])) if str(row[-1]).strip() else ""
            mes_cells.append(Paragraph(f"<b>{total_val}</b>", STYLE_MES) if total_val else "")
            data.append(mes_cells)
            mes_row_indices.append(len(data) - 1)
        elif first.startswith(MARCA_SUBTOTAL):
            sub_cells = [""] * (n_cols - 1)
            sub_cells.append(Paragraph(f"<b>{_safe(str(row[-1]))}</b>", STYLE_MES))
            data.append(sub_cells)
            sub_row_indices.append(len(data) - 1)
        else:
            data.append([_safe(str(cell)) for cell in row])

    print(f"  Tabela montada em {time.time() - t0:.1f}s")
    print(f"  {len(mes_row_indices)} separadores de mes, {len(sub_row_indices)} subtotais")

    # --- Larguras ---
    CELL_PAD = 2
    MIN_COL_W = 20
    usable_width = usable_w

    mask = ~(df_str.iloc[:, 0].str.startswith(MARCA_MES) |
             df_str.iloc[:, 0].str.startswith(MARCA_SUBTOTAL))
    df_data = df_str[mask]
    max_lens = df_data.apply(lambda c: c.map(len).max()).values
    header_lens = [len(str(c)) for c in df_str.columns]
    col_widths = [
        max(min(max(ml, hl) * 4.5, 250), MIN_COL_W)
        for ml, hl in zip(max_lens, header_lens)
    ]
    total = sum(col_widths)
    if total > usable_width:
        scale = usable_width / total
        col_widths = [max(w * scale, MIN_COL_W) for w in col_widths]

    # --- Estilos ---
    style_cmds = [
        ("GRID",       (0, 0), (-1, -1), 0.25, colors.black),
        ("BACKGROUND", (0, 0), (-1, 1),  COR_AZUL),
        ("TEXTCOLOR",  (0, 0), (-1, 1),  colors.whitesmoke),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE",   (0, 2), (-1, -1), 6.5),
        ("ALIGN",      (0, 0), (-1, 1),  "CENTER"),
        ("LEFTPADDING",   (0, 0), (-1, -1), CELL_PAD),
        ("RIGHTPADDING",  (0, 0), (-1, -1), CELL_PAD),
        ("TOPPADDING",    (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]

    # Spans do grupo (linha 0)
    col_idx = 0
    for _, span in GRUPOS:
        if span > 1 and col_idx + span - 1 < n_cols:
            style_cmds.append(
                ("SPAN", (col_idx, 0), (col_idx + span - 1, 0))
            )
        col_idx += span

    # Estilo das linhas de mês e subtotal
    special = set(mes_row_indices) | set(sub_row_indices)

    for mes_idx in mes_row_indices:
        style_cmds.append(("SPAN", (0, mes_idx), (n_cols - 2, mes_idx)))
        style_cmds.append(("BACKGROUND", (0, mes_idx), (-1, mes_idx), COR_MES))
        style_cmds.append(("ALIGN", (0, mes_idx), (-1, mes_idx), "LEFT"))
        style_cmds.append(("ALIGN", (n_cols - 1, mes_idx), (n_cols - 1, mes_idx), "RIGHT"))

    for sub_idx in sub_row_indices:
        style_cmds.append(("BACKGROUND", (0, sub_idx), (-1, sub_idx), colors.HexColor("#FFF2CC")))
        style_cmds.append(("ALIGN", (n_cols - 1, sub_idx), (n_cols - 1, sub_idx), "RIGHT"))

    # Linhas de dados alternadas (ignorando especiais)
    data_counter = 0
    for r in range(2, len(data)):
        if r in special:
            continue
        bg = colors.white if data_counter % 2 == 0 else COR_AZUL_CLARO
        style_cmds.append(("BACKGROUND", (0, r), (-1, r), bg))
        data_counter += 1

    doc = SimpleDocTemplate(
        saida,
        pagesize=landscape(A3),
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN,
    )

    table = Table(data, colWidths=col_widths, repeatRows=2)
    table.setStyle(TableStyle(style_cmds))

    elements.append(table)

    # --- Assinatura (Auditores Fiscais) ---
    assinatura = header_info.get("assinatura", [])
    if assinatura:
        elements.append(Spacer(1, 8 * mm))
        STYLE_SIGN_H = ParagraphStyle(
            "sign_h", fontSize=8, leading=10, alignment=TA_CENTER,
            fontName="Helvetica-Bold",
        )
        STYLE_SIGN_V = ParagraphStyle(
            "sign_v", fontSize=8, leading=10, alignment=TA_CENTER,
        )

        sign_cols = ["Auditores Fiscais", "Matrícula", "Assinatura", "Contribuinte"]
        sign_header = [Paragraph(f"<b>{_safe(c)}</b>", STYLE_SIGN_H) for c in sign_cols]

        sign_data = [sign_header]
        for line in assinatura:
            if any(k in str(line) for k in ("Auditores", "Matr")):
                continue
            padded = line + [""] * (4 - len(line))
            sign_data.append([Paragraph(_safe(v), STYLE_SIGN_V) for v in padded[:4]])

        sign_widths = [usable_width * 0.35, usable_width * 0.15,
                       usable_width * 0.25, usable_width * 0.25]
        sign_table = Table(sign_data, colWidths=sign_widths)
        sign_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), COR_AZUL),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(sign_table)

    print("  Gerando PDF...")
    t0 = time.time()
    doc.build(elements)
    print(f"  PDF gerado em {time.time() - t0:.1f}s")


def _pos_processar(df: pd.DataFrame) -> pd.DataFrame:
    """Corrige linhas com colunas desalinhadas."""
    fixes = 0

    for idx, row in df.iterrows():
        first = str(row.iloc[0])
        if first.startswith(MARCA_MES) or first.startswith(MARCA_SUBTOTAL):
            continue

        uf_val = str(row["UF"]).strip()
        cfop_val = str(row["CFOP"]).strip()

        # UF com valor de CFOP (4 digitos) e CFOP vazio → deslocar
        if re.match(r"^\d{4}$", uf_val) and cfop_val in ("nan", "", "NaN"):
            df.at[idx, "CFOP"] = uf_val
            df.at[idx, "UF"] = ""
            fixes += 1

        # Ct Sefaz vazio → recuperar: Ct Sefaz = Ct Contrib + Dif CT
        cs_val = row["Ct Sefaz"]
        cs_empty = pd.isna(cs_val) or str(cs_val).strip() in ("", "nan")
        if cs_empty:
            cc = row["Ct Contrib"]
            dct = row["Dif CT"]
            cc_ok = not (pd.isna(cc) or str(cc).strip() in ("", "nan"))
            dct_ok = not (pd.isna(dct) or str(dct).strip() in ("", "nan"))
            if cc_ok and dct_ok:
                try:
                    df.at[idx, "Ct Sefaz"] = float(cc) + float(dct)
                    fixes += 1
                except (ValueError, TypeError):
                    pass

    if fixes:
        print(f"  Pos-processamento: {fixes} correcoes aplicadas")
    return df


def _salvar_excel(df: pd.DataFrame, caminho: str, header_info: dict = None) -> None:
    """Salva Excel com cabeçalho institucional e formatação correta."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XlImage

    df_out = df.copy()

    # Formatar colunas
    for col in ["CNPJ/CPF", "Nota Fiscal", "Modelo", "CFOP"]:
        if col in df_out.columns:
            df_out[col] = df_out[col].astype(str).str.replace(".0", "", regex=False)
            df_out[col] = df_out[col].replace({"nan": "", "NaN": ""})

    if "Data" in df_out.columns:
        def _fmt_data(v):
            s = str(v).strip()
            if s in ("", "nan", "NaT", "None"):
                return ""
            try:
                return pd.to_datetime(s).strftime("%d/%m/%Y")
            except Exception:
                return s
        df_out["Data"] = df_out["Data"].apply(_fmt_data)

    df_out = df_out.replace({"nan": "", "NaT": "", "None": ""})

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    azul = "4472C4"
    azul_claro = "D9E2F3"
    font_header = Font(bold=True, color="FFFFFF", size=8)
    fill_azul = PatternFill(start_color=azul, end_color=azul, fill_type="solid")
    fill_azul_claro = PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type="solid")

    row_offset = 1

    if header_info:
        # Linha 1: Órgão + Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        c = ws.cell(row=1, column=1, value=header_info.get("orgao", ""))
        c.font = Font(bold=True, size=9)

        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=14)
        c = ws.cell(row=1, column=5, value=header_info.get("titulo", ""))
        c.font = Font(bold=True, italic=True, size=11, color="1F4E79")

        # Logo
        logo_path = header_info.get("logo")
        if logo_path and os.path.isfile(logo_path):
            img = XlImage(logo_path)
            img.width = 50
            img.height = 65
            ws.add_image(img, "A1")

        # Linha 2: Dados da empresa
        empresa_parts = header_info.get("empresa_parts", [])
        empresa_text = "    ".join(empresa_parts)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)
        c = ws.cell(row=2, column=1, value=empresa_text)
        c.font = Font(size=8)

        # Linha 3: Período
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=14)
        c = ws.cell(row=3, column=1, value=header_info.get("periodo", ""))
        c.font = Font(bold=True, size=8)

        # Linha 4: Grupos (Documento, Destino, Produto, Tributação)
        grupo_spans = [("Documento", 1, 3), ("Destino", 4, 7),
                       ("Produto", 8, 9), ("Tributação", 10, 14)]
        for nome, c1, c2 in grupo_spans:
            ws.merge_cells(start_row=4, start_column=c1, end_row=4, end_column=c2)
            c = ws.cell(row=4, column=c1, value=nome)
            c.font = font_header
            c.fill = fill_azul
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            for ci in range(c1 + 1, c2 + 1):
                cell = ws.cell(row=4, column=ci)
                cell.fill = fill_azul
                cell.border = border

        row_offset = 5

    # Linha de cabeçalho das colunas
    for ci, col_name in enumerate(df_out.columns, 1):
        c = ws.cell(row=row_offset, column=ci, value=col_name)
        c.font = font_header
        c.fill = fill_azul
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    # Dados
    n_cols = len(df_out.columns)
    text_cols_idx = set()
    for col in ["CNPJ/CPF", "Nota Fiscal", "Modelo", "CFOP", "Data"]:
        if col in df_out.columns:
            text_cols_idx.add(df_out.columns.get_loc(col))

    fill_mes = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_sub = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    data_counter = 0
    cur_row = row_offset + 1

    for _, row in df_out.iterrows():
        first = str(row.iloc[0]).strip()

        if first.startswith("__MES__"):
            mes_label = first.replace("__MES__", "")
            total_val = str(row.iloc[-1]).strip()
            if total_val in ("", "nan", "NaT", "None"):
                total_val = ""

            ws.merge_cells(start_row=cur_row, start_column=1,
                           end_row=cur_row, end_column=n_cols - 1)
            c = ws.cell(row=cur_row, column=1, value=mes_label)
            c.font = Font(bold=True, size=9)
            c.fill = fill_mes
            c.border = border
            for ci in range(2, n_cols):
                cell = ws.cell(row=cur_row, column=ci)
                cell.fill = fill_mes
                cell.border = border

            c = ws.cell(row=cur_row, column=n_cols)
            if total_val:
                try:
                    c.value = float(total_val)
                except ValueError:
                    c.value = total_val
            c.font = Font(bold=True, size=9)
            c.fill = fill_mes
            c.border = border
            c.alignment = Alignment(horizontal="right")
            cur_row += 1
            continue

        if first.startswith("__SUB__"):
            total_val = str(row.iloc[-1]).strip()
            for ci in range(1, n_cols):
                cell = ws.cell(row=cur_row, column=ci)
                cell.fill = fill_sub
                cell.border = border
            c = ws.cell(row=cur_row, column=n_cols)
            if total_val not in ("", "nan"):
                try:
                    c.value = float(total_val)
                except ValueError:
                    c.value = total_val
            c.font = Font(bold=True, size=8)
            c.fill = fill_sub
            c.border = border
            c.alignment = Alignment(horizontal="right")
            cur_row += 1
            continue

        for ci, val in enumerate(row):
            c = ws.cell(row=cur_row, column=ci + 1)
            s = str(val).strip()
            if s in ("", "nan", "NaT", "None"):
                c.value = ""
            elif ci in text_cols_idx:
                c.value = s
                c.number_format = "@"
            else:
                try:
                    c.value = float(s)
                except (ValueError, TypeError):
                    c.value = s

            c.font = Font(size=8)
            c.border = border
            if data_counter % 2 == 1:
                c.fill = fill_azul_claro

        data_counter += 1
        cur_row += 1

    # Assinatura
    assinatura = header_info.get("assinatura", []) if header_info else []
    if assinatura:
        cur_row += 1
        sign_cols = ["Auditores Fiscais", "Matrícula", "Assinatura", "Contribuinte"]
        col_spans = [(1, 4), (5, 7), (8, 10), (11, 14)]

        for ci, ((c1, c2), label) in enumerate(zip(col_spans, sign_cols)):
            ws.merge_cells(start_row=cur_row, start_column=c1,
                           end_row=cur_row, end_column=c2)
            c = ws.cell(row=cur_row, column=c1, value=label)
            c.font = font_header
            c.fill = fill_azul
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            for col_i in range(c1 + 1, c2 + 1):
                cell = ws.cell(row=cur_row, column=col_i)
                cell.fill = fill_azul
                cell.border = border

        cur_row += 1
        for line in assinatura:
            if any(k in str(line) for k in ("Auditores", "Matr")):
                continue
            padded = line + [""] * (4 - len(line))
            for ci, ((c1, c2), val) in enumerate(zip(col_spans, padded[:4])):
                ws.merge_cells(start_row=cur_row, start_column=c1,
                               end_row=cur_row, end_column=c2)
                c = ws.cell(row=cur_row, column=c1, value=val)
                c.font = Font(size=9)
                c.border = border
                for col_i in range(c1 + 1, c2 + 1):
                    ws.cell(row=cur_row, column=col_i).border = border
            cur_row += 1

    # Ajustar largura das colunas
    for ci in range(1, n_cols + 1):
        col_letter = get_column_letter(ci)
        max_len = len(str(df_out.columns[ci - 1]))
        for ri in range(min(100, len(df_out))):
            val = str(df_out.iloc[ri, ci - 1])
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    ws.row_dimensions[1].height = 55 if header_info else None

    wb.save(caminho)


def main():
    if len(sys.argv) > 1:
        arquivo = sys.argv[1]
    else:
        xlsx_files = [f for f in os.listdir(".") if f.endswith(".xlsx") and not f.endswith("_unificado.xlsx")]
        if len(xlsx_files) == 1:
            arquivo = xlsx_files[0]
        elif xlsx_files:
            print("Arquivos .xlsx encontrados:")
            for i, f in enumerate(xlsx_files, 1):
                print(f"  {i}. {f}")
            print(f"\nUso: python {os.path.basename(__file__)} <arquivo.xlsx>")
            sys.exit(1)
        else:
            print("Nenhum arquivo .xlsx encontrado na pasta.")
            sys.exit(1)

    if not os.path.isfile(arquivo):
        print(f"ERRO: arquivo '{arquivo}' nao encontrado.")
        sys.exit(1)

    t_total = time.time()
    nome_base = os.path.splitext(arquivo)[0]
    saida_excel = f"{nome_base}_unificado.xlsx"
    saida_pdf = f"{nome_base}.pdf"

    print(f"Lendo '{arquivo}'...")

    header_info = extrair_header(arquivo)
    print(f"  Titulo: {header_info.get('titulo', '-')}")
    print(f"  {header_info.get('periodo', '-')}")

    df_final = ler_dados(arquivo)
    print(f"  -> {len(df_final)} linhas | {len(df_final.columns)} colunas")

    df_final = _pos_processar(df_final)

    xls = pd.ExcelFile(arquivo)
    if len(xls.sheet_names) > 1:
        print("Salvando Excel unificado...")
        _salvar_excel(df_final, saida_excel, header_info)
        mask = ~(df_final.iloc[:, 0].astype(str).str.startswith(MARCA_MES) |
                 df_final.iloc[:, 0].astype(str).str.startswith(MARCA_SUBTOTAL))
        df_excel = df_final[mask].copy()
        print(f"  {saida_excel} salvo ({len(df_excel)} linhas de dados).")
    else:
        print("  Apenas 1 aba encontrada, pulando unificacao do Excel.")

    gerar_pdf(df_final, saida_pdf, header_info)
    print(f"  {saida_pdf} salvo.")

    verificar(df_excel if len(xls.sheet_names) > 1 else df_final)

    print(f"Concluido em {time.time() - t_total:.1f}s")


def verificar(df: pd.DataFrame) -> None:
    print()
    print("=" * 50)
    print("  VERIFICACAO DE INTEGRIDADE")
    print("=" * 50)

    total = len(df)
    erros = 0

    # Nota Fiscal
    nf = df["Nota Fiscal"].astype(str)
    nf_bad = nf[~nf.str.match(r"^\d+$")]
    if len(nf_bad):
        print(f"  [!] Nota Fiscal nao numerica: {len(nf_bad)}")
        erros += len(nf_bad)

    # Modelo
    mod = df["Modelo"].astype(str).str.replace(".0", "", regex=False)
    mod_bad = mod[~mod.isin(["55", "65", "", "nan"])]
    if len(mod_bad):
        print(f"  [!] Modelo diferente de 55/65: {len(mod_bad)}")
        erros += len(mod_bad)

    # UF
    UFS = {"GO","SP","MG","RJ","BA","PR","SC","RS","CE","PE","PA","MA",
           "MT","MS","DF","ES","PB","RN","AL","PI","SE","TO","RO","AC",
           "AP","AM","RR","","nan"}
    uf = df["UF"].astype(str)
    uf_cfop = uf[uf.str.match(r"^\d{4}$")]
    uf_ok = uf[uf.isin(UFS)]
    uf_empty = len(uf) - len(uf_ok) - len(uf_cfop)
    print(f"  UF valida: {len(uf_ok)} | vazia: {uf_empty} | com CFOP (erro): {len(uf_cfop)}")
    erros += len(uf_cfop)

    # CFOP
    cfop = df["CFOP"].astype(str).str.replace(".0", "", regex=False)
    cfop_bad = cfop[~cfop.str.match(r"^\d{4}$") & ~cfop.isin(["", "nan"])]
    if len(cfop_bad):
        print(f"  [!] CFOP invalido: {len(cfop_bad)}")
        erros += len(cfop_bad)

    # CCE
    cce = df["CCE"].astype(str)
    cce_bad = cce[~cce.isin(["Sim", "Não", "Nao", "N\xe3o", "", "nan"])]
    if len(cce_bad):
        print(f"  [!] CCE invalido: {len(cce_bad)}")
        erros += len(cce_bad)

    # Dif. Icms
    dif = df["Dif. Icms"]
    dif_blank = dif[dif.isna() | (dif.astype(str).str.strip() == "")]
    print(f"  Dif. Icms preenchido: {total - len(dif_blank)} | vazio: {len(dif_blank)}")
    erros += len(dif_blank)

    # Valor Produto
    vp = df["Valor Produto"]
    vp_blank = vp[vp.isna() | (vp.astype(str).str.strip() == "")]
    print(f"  Valor Produto preenchido: {total - len(vp_blank)} | vazio: {len(vp_blank)}")
    erros += len(vp_blank)

    # Ct Sefaz
    cs = df["Ct Sefaz"]
    cs_blank = cs[cs.isna() | (cs.astype(str).str.strip() == "")]
    print(f"  Ct Sefaz preenchido: {total - len(cs_blank)} | vazio: {len(cs_blank)}")

    # CNPJ/CPF
    cnpj = df["CNPJ/CPF"].astype(str)
    cnpj_zero = cnpj[cnpj == "0"]
    cnpj_ok = cnpj[cnpj.str.len() >= 5]
    print(f"  CNPJ/CPF valido: {len(cnpj_ok)} | zerado: {len(cnpj_zero)} | outros: {total - len(cnpj_ok) - len(cnpj_zero)}")

    print()
    if erros == 0:
        print(f"  RESULTADO: {total} linhas verificadas - TUDO OK")
    else:
        print(f"  RESULTADO: {total} linhas verificadas - {erros} alerta(s)")
    print("=" * 50)


if __name__ == "__main__":
    main()
